import datetime
import json

import pandas as pd
from openpyxl import load_workbook



class AdviceRow:
	def __init__ (self):
		self._sheet = None
		self._start = { 'column':'A', 'cell': ''}
		self._end = { 'column':'B', 'cell': ''}
		self._message = { 'column':'C', 'cell': ''}
		self._level = { 'column':'D', 'cell': ''}

	def set_index (self, index):
		for attr in [self._start, self._end, self._message, self._level]:
			attr['cell'] = '{}{}'.format (attr['column'], index)

	def set_sheet (self, sheet):
		self._sheet = sheet

	def get_start (self):
		if self._sheet:
			return self._sheet[self._start['cell']].value
		return None

	def get_string_start (self):
		return self.get_start ().strftime ('%H:%M')

	def get_end (self):
		if self._sheet:
			return self._sheet[self._end['cell']].value
		return None

	def get_string_end (self):
		return self.get_end ().strftime ('%H:%M')

	def get_message (self):
		if self._sheet:
			return self._sheet[self._message['cell']].value
		return ''

	def get_level (self):
		if self._sheet:
			return self._sheet[self._level['cell']].value
		return ''


class Advice:
	def __init__ (self, filepath):
		self._workbook = load_workbook (filename=filepath)
		self._sheet_names = ['consejos_nublado', 'consejos_soleado']

	def time_in_range (self, start, end, hour):
		return start <= hour <= end

	def get_advices (self, day, start_hour, end_hour):
		sheet = self._workbook[self._sheet_names[day]]

		start_time = datetime.time (start_hour, 0, 0)
		end_time = datetime.time (end_hour - 1, 59, 59)

		advices = []

		row = AdviceRow ()
		row.set_sheet (sheet)

		for index in range (2, sheet.max_row + 1):
			row.set_index (index)
			if (None == row.get_start () or None == row.get_end ()):
				break

			included_start = self.time_in_range (row.get_start (), 
												 row.get_end (), 
												 start_time)
			included_end = self.time_in_range (row.get_start (),
											   row.get_end (),
											   end_time)

			if included_start or included_end:
				advices.append ({
					'start': row.get_string_start (),
					'end': row.get_string_end (),
					'message': row.get_message (),
					'level': row.get_level ()
				})

		return advices


class CostApi:
	def __init__ (self, df, registers_per_hour):
		self._df = df
		self._interval = registers_per_hour
		self._columns = [
			'Ahorro (€)',
			'Coste (€)', 
			'CO2 (g)'
		]

	def get_stats (self, start_hour, end_hour):
		from_index = start_hour * self._interval
		to_index = end_hour * self._interval

		result = self._get_summations (from_index, to_index)

		return result

	def _get_summations (self, from_index, to_index):
		summation = self._df[self._columns].iloc[0:to_index].cumsum ()
		
		result = summation.loc[from_index:to_index]

		rename_columns = {
			self._columns[0]: 'ahorro_sum',
			self._columns[1]: 'coste_sum',
			self._columns[2]: 'co2_sum'
		}

		result = result.rename (columns=rename_columns)

		return result


class Housing:
	def __init__ (self, df, registers_per_hour):
		self._df = df
		self._interval = registers_per_hour
		self._columns = [
			'Consumo (W)', 
			'Generacion (W)', 
			'Autoconsumo (W)'
		]

	def get_stats (self, start_hour, end_hour):
		from_index = start_hour * self._interval
		to_index = end_hour * self._interval

		puntual = self._get_consumptions (from_index, to_index)
		summation = self._get_summations (from_index, to_index)

		result = pd.merge (puntual, 
						   summation,
						   how='outer', 
						   left_index=True, 
						   right_index=True)

		return result

	def _get_consumptions (self, from_index, to_index):
		result = self._df[[self._columns[0],
						   self._columns[1],
						   self._columns[2]]].iloc[from_index:to_index]

		result['consumo'] = result[self._columns[0]] - result[self._columns[1]]

		rename_columns = {
			self._columns[0]: 'consumo_total',
			self._columns[1]: 'generacion',
			self._columns[2]: 'autoconsumo'
		}

		result = result.rename (columns=rename_columns)

		return result

	def _get_summations (self, from_index, to_index):
		summation = self._df[[self._columns[0],
						      self._columns[2]]].iloc[0:to_index].cumsum ()

		summation['autoconsumo_percent'] = (summation[self._columns[2]] / 
											summation[self._columns[0]])
		result = summation.loc[from_index:to_index]

		rename_columns = {
			self._columns[0]: 'consumo_sum',
			self._columns[2]: 'autoconsumo_sum'
		}

		result = result.rename (columns=rename_columns)

		return result['autoconsumo_percent']


class HomeAppliance:
	def __init__ (self, df, registers_per_hour):
		self._df = df
		self._interval = registers_per_hour
		self._columns = [
			'Frigorífico (%)',
			'Lavadora (%)',
			'Vitrocerámica (%)',
			'Termo eléctrico (%)',
			'Frigorífico (W)',
			'Lavadora (W)',
			'Vitrocerámica (W)',
			'Termo eléctrico (W)'
		]

	def get_stats (self, start_hour, end_hour):
		from_index = start_hour * self._interval
		to_index = end_hour * self._interval

		result = self._get_consumptions (from_index, to_index)

		return result

	def _get_consumptions (self, from_index, to_index):
		consumptions = self._df[self._columns].iloc[from_index:to_index]

		rename_columns = {
		  	self._columns[0]: 'frigorifico_percent',
		  	self._columns[1]: 'lavadora_percent',
		  	self._columns[2]: 'vitroceramica_percent',
			self._columns[3]: 'termoelectrico_percent',
			self._columns[4]: 'frigorifico',
			self._columns[5]: 'lavadora',
			self._columns[6]: 'vitroceramica',
			self._columns[7]: 'termoelectrico'
		}

		result = consumptions.rename (columns=rename_columns)

		return result